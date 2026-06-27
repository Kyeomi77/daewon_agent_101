"""Step 4: 시험처방 도출표 .xlsx 자동 생성.

Gemini 가 추론한 최적 처방 3종을 연구원이 실험에 활용할 수 있는
Excel 처방 도출표로 출력한다.
"""
from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..schemas import FormulationProposal

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1E4D8C")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_RISK_FILL = {
    "Low": PatternFill("solid", fgColor="C6EFCE"),
    "Med": PatternFill("solid", fgColor="FFEB9C"),
    "High": PatternFill("solid", fgColor="FFC7CE"),
}
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def write_formulation_xlsx(
    proposal: FormulationProposal, out_path: str | Path
) -> Path:
    out_path = Path(out_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "시험처방 도출표"

    # 제목
    ws["A1"] = f"시험처방 도출표 — {proposal.api_name}"
    ws["A1"].font = Font(size=14, bold=True)
    bcs = proposal.bcs_class.value if proposal.bcs_class else "미상"
    ws["A2"] = f"BCS 분류: {bcs}    참조 사례: {', '.join(proposal.cited_cases) or '-'}"
    ws["A2"].font = Font(italic=True, color="595959")

    headers = ["처방 ID", "부형제", "역할", "함량(mg)", "비율(%)",
               "기대 용출률(%)", "기대 함량(%)", "리스크", "설계 근거"]
    start_row = 4
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    row = start_row + 1
    for f in proposal.formulations:
        first_row = row
        for e in f.excipients:
            ws.cell(row=row, column=1, value=f.formulation_id).border = _BORDER
            ws.cell(row=row, column=2, value=e.name).border = _BORDER
            ws.cell(row=row, column=3, value=e.function).border = _BORDER
            ws.cell(row=row, column=4, value=e.amount_mg).border = _BORDER
            ws.cell(row=row, column=5, value=e.percent).border = _BORDER
            row += 1
        last_row = row - 1
        # 처방 단위로 병합 (기대값/리스크/근거)
        if last_row >= first_row:
            for col in (1, 6, 7, 8, 9):
                ws.merge_cells(start_row=first_row, start_column=col,
                               end_row=last_row, end_column=col)
            ws.cell(row=first_row, column=6, value=f.expected_dissolution)
            ws.cell(row=first_row, column=7, value=f.expected_assay)
            risk_cell = ws.cell(row=first_row, column=8, value=f.risk_level.value)
            risk_cell.fill = _RISK_FILL.get(f.risk_level.value, _RISK_FILL["Med"])
            risk_cell.alignment = Alignment(horizontal="center", vertical="center")
            rationale_cell = ws.cell(row=first_row, column=9, value=f.rationale)
            rationale_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 열 너비
    widths = [10, 20, 14, 12, 10, 14, 14, 10, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    logger.info("처방 도출표 저장: %s", out_path)
    return out_path
